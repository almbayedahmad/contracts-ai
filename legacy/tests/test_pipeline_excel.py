import importlib, types, io
import pandas as pd
from openpyxl import load_workbook
from tests.helpers import FakeUpload

def test_run_pipeline_generates_excel(monkeypatch):
    import sys, types
    # Ensure no real extractors run
    ui = importlib.import_module('app.ui_streamlit')
    ui.REGISTRY['extractors'] = []
    # Stub summarize functions
    ui.summarize = lambda df: 'summary en'
    ui.summarize_de = lambda df: 'Zusammenfassung de'
    # Stub pipeline.postprocess.build_price_schedule
    dummy_pipeline = types.ModuleType('pipeline')
    dummy_post = types.ModuleType('pipeline.postprocess')
    def _bps(df):
        import pandas as pd
        return pd.DataFrame([{'item':'base','price':'0'}])
    dummy_post.build_price_schedule = _bps
    sys.modules['pipeline'] = dummy_pipeline
    sys.modules['pipeline.postprocess'] = dummy_post

    ui = importlib.import_module("app.ui_streamlit")

    # Stub build_entities_links to return minimal consistent frames
    def _stub_entities_links(df_spans, doc_id, full_text):
        ents = pd.DataFrame([
            {"entity_id": "E1", "name": "Partei A", "type": "party"},
            {"entity_id": "E2", "name": "Partei B", "type": "party"},
        ])
        links = pd.DataFrame([
            {"from_id": "E1", "to_id": "E2", "type": "relates_to"}
        ])
        return ents, links

    # Stub price schedule (tables unused)
    def _stub_price_from_tables(tables):
        return pd.DataFrame([{"item": "base", "price": "0"}])

    # Make sure compliance evaluation does nothing harmful
    dummy_rules = types.ModuleType("rules")
    dummy_engine = types.ModuleType("rules.engine")
    def _eval_compliance(frames, policies):
        return pd.DataFrame([])
    dummy_engine.evaluate_compliance = _eval_compliance

    # Inject stubs into sys.modules and module namespace
    import sys
    sys.modules["rules"] = dummy_rules
    sys.modules["rules.engine"] = dummy_engine

    ui.build_entities_links = _stub_entities_links
    ui.build_price_schedule_from_tables = _stub_price_from_tables

    # Minimal text (.txt) to avoid docx/pdf deps in the function path
    text = b"Dies ist ein Testvertrag. Gerichtsstand Berlin. deutsches Recht."
    up = FakeUpload("mini.txt", text)

    text_out, spans, comp, ents, links, summary_de, excel_buf = ui.run_local_pipeline(up)
    assert isinstance(excel_buf, io.BytesIO)
    excel_buf.seek(0)
    wb = load_workbook(excel_buf, read_only=True, data_only=True)
    sheets = set(wb.sheetnames)
    required = {"Spans","Entities","Links","Compliance","PriceSchedule","ContractSummary","ContractSummary_DE"}
    assert required.issubset(sheets), f"Missing sheets: {required - sheets}"